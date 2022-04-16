import os
import openpyxl

with open(r'D:\CODE\my_nas\py\duoduomaicai_scripts\source_file\maicai_orders.txt', 'r', encoding='utf-8') as f:
    lines = f.readlines()

wb = openpyxl.load_workbook(r'D:\CODE\my_nas\py\duoduomaicai_scripts\source_file\4-17.xlsx')

wb.create_sheet(index=len(wb.worksheets), title='商品总表')

goods_details_dict = {}
goods_name_set = set()
building_name_set = set()
goods_all_count_dict = {}

building_2_info_dict = {}
for line in lines:
    line = line.strip().split('\t')
    if len(line) != 6:
        print(line)
    user_name, wechat_name, phone_number, goods_name, goods_count, address = line
    goods_name_set.add(goods_name)
    goods_all_count_dict[goods_name] = goods_all_count_dict.get(goods_name, 0) + int(goods_count)
    building_number = address.split('号')[0] + '号'
    building_name_set.add(building_number)

    if building_number not in building_2_info_dict:
        building_2_info_dict[building_number] = []
    building_2_info_dict[building_number].append([user_name, wechat_name, phone_number, goods_name, goods_count, address])

for i in list(building_name_set):
    for j in list(goods_name_set):
        if i not in goods_details_dict:
            goods_details_dict[i] = {}
        goods_details_dict[i][j] = 0

building_name_list = list(sorted(list(building_name_set)))
goods_name_list = list(sorted(list(goods_name_set)))

# for building_number, v in building_2_info_dict.items():
#     print(building_number, len(v))
def process(wb, building_number, building_details):
    sheet_id = len(wb.worksheets)
    wb.create_sheet(index=len(wb.worksheets), title=building_number)
    sheet = wb.worksheets[sheet_id]
    row, col = 1, 1
    # sheet.cell(row, col).value = '1'
    # f = open(os.path.join('maicai', f'{building_number}.txt'), 'w', encoding='utf-8')
    label_bar = ['用户名', '微信名', '电话号', '商品名', '商品数', '地址']

    res = []
    res.append(label_bar)

    # f.write(f'{label_bar}\n')
    for building_detail in building_details:
        # tmp_detail = '\t'.join(building_detail)
        # f.write(f'{tmp_detail}\n')
        res += building_detail
    # f.write('\n' * 5)
    res.append([''] * 6)

    for building_detail in building_details:
        user_name, wechat_name, phone_number, goods_name, goods_count, address = building_detail
        goods_details_dict[building_number][goods_name] += int(goods_count)
    # f.write(f'商品统计表\n')
    res.append(['商品统计表'])
    sort_goods_details = list(goods_details_dict[building_number].items())
    sort_goods_details.sort()
    for v in sort_goods_details:
        # f.write(f'{v[0]}\t{v[1]}\n')
        res.append([v[0], v[1]])

    for info_list in res:
        col = 1
        for uninsert_data in info_list:
            try:
                sheet.cell(row, col).value = uninsert_data
                col += 1
            except:
                print()
        row += 1
    # f.close()


for building_number in building_name_list:
    process(wb, building_number, building_2_info_dict[building_number])

# with open(os.path.join('maicai', '总表.txt'), 'w', encoding='utf-8') as f:
res = []
label_bar = ['商品名', '总表累加', '分表累加'] + building_name_list
res.append(label_bar)
# label_bar = '\t'.join(label_bar)
# f.write(f'{label_bar}\n')
for goods_name in goods_name_list:
    colume = []
    colume.append(goods_name)
    colume.append(goods_all_count_dict[goods_name])
    tmp_colume = []
    for building_name in building_name_list:
        tmp_count = goods_details_dict[building_name][goods_name]
        tmp_colume.append(tmp_count)
    colume.append(sum(tmp_colume))
    colume += tmp_colume
    colume = list(map(str, colume))
    res.append(colume)
    # colume = '\t'.join(colume)
    # f.write(f'{colume}\n')
sheet = wb.worksheets[1]
row, col = 1, 1
for info_list in res:
    col = 1
    for uninsert_data in info_list:
        try:
            sheet.cell(row, col).value = uninsert_data
            col += 1
        except:
            print()
    row += 1

wb.save(r'D:\CODE\my_nas\py\duoduomaicai_scripts\test\other_order_info.txt')